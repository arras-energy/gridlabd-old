"""Create NAICS data file from MECS and NERC data files

Run command:

	host% python3 mecs.py
	Loaded table 9.1 ok
	Loaded table 11.1 ok
	Loaded table 5.1 ok
	Saved ../mecs_data.csv ok
	Loaded ../nerc_data.csv ok
	Merged ok
	Saved ../naics_data_file.csv ok
	
"""
from openpyxl import load_workbook
from pandas import DataFrame, concat, read_csv

import config

table5 = load_workbook(filename="MECS2014_Table_5-1.xlsx")["Table 5.1"]
table9 = load_workbook(filename="MECS2014_Table_9-1.xlsx")["Table 9.1"]
table11 = load_workbook(filename="MECS2014_Table_11-1.xlsx")["Table 11.1"]

columns = [
	"NAICS", "DESCRIPTION", "REGION", # index
	"FLOORAREA","SITES","FLOORAREA_PER_SITE","BUILDINGS","BUILDINGS_PER_SITE", # Table 9
	"PURCHASES","TRANSFERS","GENERATION","SALES","NET_DEMAND", # Table 11
	]

#
# Table 9 - Buildings
#
data = []
for values in table9.iter_rows(min_row=18,min_col=1,max_row=98,max_col=7,values_only=True):
	row = [int(values[0])]
	row.extend(values[1:])
	data.append(row)
mecs9 = DataFrame.from_records(data,index=["NAICS"],
			columns = ["NAICS","DESCRIPTION","FLOORAREA[Msf]","SITES[unit]","FLOORAREA_PER_SITE[sf/unit]","BUILDINGS[unit]","BUILDINGS_PER_SITE[unit/unit]"])
valid_codes = mecs9.index
print("Loaded table 9.1 ok")

#
# Table 11 - Regional electricity use
#
regions = {
	"NERC":(13,93),
	"NORTHEAST":(97,177),
	"MIDWEST":(181,261),
	"SOUTH":(265,345),
	"WEST":(349,429),
	}
data = []
for region, rows in regions.items():
		for values in table11.iter_rows(min_row=rows[0],min_col=1,max_row=rows[1],max_col=7,values_only=True):
			try:
				index = int(values[0].strip())
				if index in valid_codes:
					row = [index,region]
					for value in values[2:]:
						if value == "W": # withheld due to low sample count
							value = float('nan')
						elif value == "Q": # withheld due to high RSE
							value = float('nan')
						elif value == "*": # value < 0.5
							value = 0.0
						elif value == "NA": # value not available
							value = float('nan')
						row.append(value)
					data.append(row)
			except:
				pass
mecs11 = DataFrame.from_records(data,index=["NAICS"],
			columns=["NAICS","REGION","PURCHASES[GWh]","TRANSFERS[GWh]","GENERATION[GWh]","SALES[GWh]","NET_DEMAND[GWh]"])
print("Loaded table 11.1 ok")

#
# Table 5 - End-use decomposition
#
data = []
index = None
electrification = [None,None,None,None]
for values in table5.iter_rows(min_row=17,min_col=1,max_row=1819,max_col=4,values_only=True):
	if values[0]:
		try:
			if index:
				row = [index]
				row.extend(electrification)
				for enduse in ["A","B","C","D","E","I","Z"]:
					row.append(enduse_composition[enduse]/enduse_composition["TOTAL"])
				data.append(row)
			index = int(values[0])
			if not index in valid_codes:
				index = None
		except:
			continue
		enduse_composition = {"A":0.0,"B":0.0,"C":0.0,"D":0.0,"E":0.0,"I":0.0,"Z":0.0,"TOTAL":0.0}
	elif not values[1]:
		continue
	tag = values[1].strip()
	if index and values[2] and values[3]:
		try:
			electric = float(values[3])
			if tag == "TOTAL FUEL CONSUMPTION":
				total = float(values[2])
				electrification[0] = electric
				electrification[1] = total/electric*3.412
			elif tag == "End Use Not Reported":
				total = float(values[2])
				electrification[2] = electric
				electrification[3] = total/electric*3.412
				components = config.enduses[tag]
			elif tag in config.enduses.keys():
				components = config.enduses[tag]
			if components:
				for enduse, value in components.items():
					enduse_composition[enduse] += value
					enduse_composition["TOTAL"] += value
				components = None				
		except:
			pass
if index:
	row = [index]
	row.extend(electrification)
	data.append(row)
mecs5 = DataFrame.from_records(data,index=["NAICS"],
			columns=["NAICS","REPORTED_ELECTRIC[GWh]","REPORTED_ELECTRIFICATION[pu]","UNREPORTED_ELECTRIC[GWh]","UNREPORTED_ELECTRIFICATION[pu]",
			"A","B","C","D","E","I","Z"])
print("Loaded table 5.1 ok")

mecs = mecs9.join(mecs11).join(mecs5)
local = mecs["REGION"]!="NERC"
for column in ["FLOORAREA[Msf]","BUILDINGS[unit]","SITES[unit]"]:
	mecs.at[local,column] = float('nan')

mecs.index.name = "NAICS"
mecs.reset_index(inplace=True)
mecs.set_index(["NAICS","REGION"],inplace=True)

mecs.to_csv(config.mecs_csv)
print(f"Saved {config.mecs_csv} ok")

nerc = read_csv(config.nerc_csv,converters={"NAICS":int}).set_index("Code")
print(f"Loaded {config.nerc_csv} ok")
for row in mecs.iterrows():
	name = f"MECS_{row[0][1]}_{row[0][0]}"
	nerc.loc[name,"NAICS"] = int(row[0][0])
	for item, value in dict(
			DESCRIPTION="Industrial Load Type",
			A="IA",
			B="IB",
			C="IC",
			D="MD",
			E="PwrEl",
			I="I",
			Z="Z").items():
		if type(row[1][item]) is float:
			nerc.loc[name,value] = round(row[1][item],3)
		elif type(row[1][item]) is str and len(row[1][item]) > 0:
			nerc.loc[name,value] = row[1][item].replace(',','').strip()
		else:
			nerc.loc[name,value] = "NA"
print(f"Merged ok")

nerc.reset_index().set_index(["NAICS","Code"]).to_csv(config.naics_csv,float_format="%g")
print(f"Saved {config.naics_csv} ok")
